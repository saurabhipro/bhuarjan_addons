import os
import re
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -------------------------
# Professional Styling Constants
# -------------------------
NAVY_FILL = PatternFill("solid", fgColor="1F4E78")  # Professional Navy
WHITE_TEXT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")  # Light Blue for headers
STRIPE_FILL = PatternFill("solid", fgColor="F2F2F2")  # Very Light Gray for zebra stripes
THIN_BORDER = Border(
    left=Side(style="thin", color="B2B2B2"),
    right=Side(style="thin", color="B2B2B2"),
    top=Side(style="thin", color="B2B2B2"),
    bottom=Side(style="thin", color="B2B2B2")
)

# -------------------------
# Helpers
# -------------------------
def _safe_filename(name: str, fallback: str = "tender") -> str:
    s = str(name or "").strip()
    if not s: return fallback
    s = re.sub(r"[\\/:*?\"<>|]+", "_", s)
    s = re.sub(r"\s+", "_", s).strip("_")
    return s or fallback

def _safe_int(v, default=0):
    try:
        return int(v)
    except:
        return default

def _apply_header_style(cell):
    cell.fill = NAVY_FILL
    cell.font = WHITE_TEXT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

def _add_sheet_summary(ws, tender_data, title, col_count=6):
    """Adds a consistent blue header summary to the top of every sheet."""
    # Main Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    top_cell = ws.cell(row=1, column=1, value=f"{title} - TENDER REPORT")
    _apply_header_style(top_cell)
    top_cell.font = Font(color="FFFFFF", bold=True, size=14)
    
    # Quick Info Row
    t_id = (tender_data or {}).get("tenderId", "N/A")
    ref = (tender_data or {}).get("refNo", "N/A")
    
    ws.cell(row=2, column=1, value="Tender ID:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=t_id)
    ws.cell(row=2, column=3, value="Ref No:").font = Font(bold=True)
    ws.cell(row=2, column=4, value=ref)
    
    for r in range(1, 3):
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

# -------------------------
# Main Export
# -------------------------
def export_tender_excel(
    data: dict,
    bidders: list = None,
    payments_by_company: list = None,
    work_experience_by_company: list = None,
    analytics: dict = None,
) -> str:
    out_dir = os.path.join(str(settings.BASE_DIR), "excels")
    os.makedirs(out_dir, exist_ok=True)

    raw_tender_id = (data or {}).get("tenderId") or "tender"
    tender_id = _safe_filename(raw_tender_id, "tender")
    out_path = os.path.join(out_dir, f"{tender_id}.xlsx")

    wb = Workbook()

    # 1. General Info
    ws = wb.active
    ws.title = "General Information"
    _setup_general_info(ws, data)

    # 2. Bidders
    if bidders:
        add_bidder_details_sheet(wb, data, bidders)

    # 3. Payments
    if payments_by_company:
        add_payment_details_sheet(wb, data, payments_by_company)

    # 4. Work Experience
    if work_experience_by_company:
        add_work_experience_sheet(wb, data, work_experience_by_company)

    # 5. Eligibility
    add_eligibility_criteria_sheet(wb, data)

    # 6. Analytics
    add_analytics_sheet(wb, data, analytics=analytics)

    wb.save(out_path)
    return out_path


def _setup_general_info(ws, data):
    _add_sheet_summary(ws, data, "GENERAL INFORMATION", col_count=6)
    
    col_widths = [25, 35, 25, 35, 25, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = [
        ("Tender ID", "tenderId", "Ref. No.", "refNo", "Tender Creator", "tenderCreator"),
        ("Category", "procurementCategory", "Tender Type", "tenderType", "Hierarchy", "organizationHierarchy"),
        ("Value (INR)", "estimatedValueINR", "Currency", "tenderCurrency", "Bidding", "biddingCurrency"),
        ("Validity (Days)", "offerValidityDays", "Prev Tender", "previousTenderNo", "Published", "publishedOn"),
        ("Start Date", "bidSubmissionStart", "End Date", "bidSubmissionEnd", "Opened On", "tenderOpenedOn"),
    ]

    curr_row = 4
    for label_set in rows:
        for i in range(0, 6, 2):
            # Label
            lbl = ws.cell(row=curr_row, column=i+1, value=label_set[i])
            lbl.fill = HEADER_FILL
            lbl.font = Font(bold=True)
            # Value
            val = ws.cell(row=curr_row, column=i+2, value=(data or {}).get(label_set[i+1], ""))
            
            lbl.border = val.border = THIN_BORDER
        curr_row += 1

    # Large text areas
    for field in [("Description", "description"), ("NIT", "nit")]:
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=1)
        lbl = ws.cell(row=curr_row, column=1, value=field[0])
        lbl.fill = HEADER_FILL
        lbl.font = Font(bold=True)
        lbl.border = THIN_BORDER
        
        ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=6)
        val = ws.cell(row=curr_row, column=2, value=(data or {}).get(field[1], ""))
        val.alignment = Alignment(wrap_text=True)
        for c in range(2, 7): ws.cell(row=curr_row, column=c).border = THIN_BORDER
        curr_row += 1


def add_bidder_details_sheet(wb, tender_data, bidders):
    ws = wb.create_sheet("Bidder Details")
    _add_sheet_summary(ws, tender_data, "BIDDER LIST", col_count=9)
    
    headers = ["Company Name", "Address", "Email", "Contact Person", "Phone", "PAN", "GSTIN", "Registration", "Validity"]
    start_row = 4
    
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        _apply_header_style(c)
        ws.column_dimensions[get_column_letter(i)].width = 25

    row_num = start_row + 1
    for i, b in enumerate(bidders):
        vals = [b.get("vendorCompanyName"), b.get("companyAddress"), b.get("emailId"), b.get("contactPerson"), 
                b.get("contactNo"), b.get("pan"), b.get("gstin"), b.get("placeOfRegistration"), b.get("offerValidityDays")]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
            if i % 2 == 1: cell.fill = STRIPE_FILL
        row_num += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{row_num-1}"


def add_payment_details_sheet(wb, tender_data, payments_by_company):
    ws = wb.create_sheet("Payment Details")
    _add_sheet_summary(ws, tender_data, "PAYMENT TRACKER", col_count=8)
    
    headers = ["Company Name", "Vendor", "Mode", "Bank", "Transaction ID", "Amount (INR)", "Date", "Status"]
    start_row = 4
    for i, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=start_row, column=i, value=h))
        ws.column_dimensions[get_column_letter(i)].width = 22

    row_num = start_row + 1
    for i, group in enumerate(payments_by_company):
        comp = group.get("companyName", "")
        for p in group.get("payments", []):
            vals = [comp, p.get("vendor"), p.get("paymentMode"), p.get("bankName"), p.get("transactionId"), p.get("amountINR"), p.get("transactionDate"), p.get("status")]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = THIN_BORDER
                if i % 2 == 1: cell.fill = STRIPE_FILL
            row_num += 1
    ws.freeze_panes = "A5"


def add_work_experience_sheet(wb, tender_data, work_experience_by_company):
    ws = wb.create_sheet("Work Experience")
    _add_sheet_summary(ws, tender_data, "EXPERIENCE LOG", col_count=9)
    
    headers = ["Vendor", "Project Name", "Employer", "Location", "Amount", "Start", "End", "Certificate", "Link"]
    start_row = 4
    for i, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=start_row, column=i, value=h))
        ws.column_dimensions[get_column_letter(i)].width = 25

    row_num = start_row + 1
    for i, group in enumerate(work_experience_by_company):
        for w in group.get("workExperience", []):
            vals = [w.get("vendorCompanyName"), w.get("nameOfWork"), w.get("employer"), w.get("location"), w.get("contractAmountINR"), w.get("dateOfStart"), w.get("dateOfCompletion"), w.get("completionCertificate"), w.get("attachment")]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = THIN_BORDER
                if i % 2 == 1: cell.fill = STRIPE_FILL
            row_num += 1
    ws.freeze_panes = "A5"


def add_eligibility_criteria_sheet(wb, tender_data):
    ws = wb.create_sheet("Eligibility Criteria")
    _add_sheet_summary(ws, tender_data, "QUALIFICATION CRITERIA", col_count=3)
    
    headers = ["Sl.", "Required Criteria", "Supporting Document Required"]
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 65

    start_row = 4
    for i, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=start_row, column=i, value=h))

    criteria = (tender_data or {}).get("bidderEligibilityCriteria") or []
    row_num = start_row + 1
    
    if not criteria:
        ws.cell(row=row_num, column=2, value="No criteria extracted from document.").font = Font(italic=True)
    else:
        for i, r in enumerate(criteria):
            ws.cell(row=row_num, column=1, value=r.get("slNo")).border = THIN_BORDER
            c2 = ws.cell(row=row_num, column=2, value=r.get("criteria"))
            c3 = ws.cell(row=row_num, column=3, value=r.get("supportingDocument"))
            c2.alignment = c3.alignment = Alignment(wrap_text=True, vertical="top")
            c2.border = c3.border = THIN_BORDER
            if i % 2 == 1: 
                ws.cell(row=row_num, column=1).fill = STRIPE_FILL
                c2.fill = c3.fill = STRIPE_FILL
            row_num += 1


def add_analytics_sheet(wb, tender_data, analytics=None):
    ws = wb.create_sheet("Processing Analytics")
    _add_sheet_summary(ws, tender_data, "SYSTEM PERFORMANCE", col_count=5)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    
    # Simple key-value display for performance
    curr_row = 4
    stats_header = ws.cell(row=curr_row, column=1, value="Metric Name")
    val_header = ws.cell(row=curr_row, column=2, value="Value")
    _apply_header_style(stats_header)
    _apply_header_style(val_header)
    
    curr_row += 1
    t_an = (tender_data or {}).get("tenderAnalytics") or {}
    metrics = [
        ("AI Model Used", t_an.get("model")),
        ("Extraction Time (ms)", _safe_int(t_an.get("durationMs"))),
        ("Total Tokens Consumed", _safe_int((t_an.get("tokens") or {}).get("totalTokens")))
    ]
    
    if analytics:
        metrics.extend([
            ("Total Processing Time (s)", analytics.get("durationSeconds")),
            ("Total Companies Found", analytics.get("companiesDetected")),
            ("Gemini API Calls", analytics.get("geminiCallsTotal"))
        ])

    for label, val in metrics:
        l_cell = ws.cell(row=curr_row, column=1, value=label)
        v_cell = ws.cell(row=curr_row, column=2, value=val)
        l_cell.border = v_cell.border = THIN_BORDER
        l_cell.fill = HEADER_FILL
        curr_row += 1